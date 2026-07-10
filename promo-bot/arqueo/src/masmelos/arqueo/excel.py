"""Arma el Excel de arqueo que revisa el tesorero (y gerencia).

Formato pensado para imprimirse: Arial, encabezados en gris claro, negativos
entre paréntesis, sin colores semánticos (limitación de impresoras — regla
del proyecto). La severidad se comunica con texto (ROJA/AMARILLA/INFO) y
negrita, no con rojo/amarillo.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_FONT = Font(name="Arial", size=10)
_FONT_BOLD = Font(name="Arial", size=10, bold=True)
_FONT_TITULO = Font(name="Arial", size=14, bold=True)
_FILL_HEADER = PatternFill("solid", start_color="D9D9D9")
_BORDE = Border(bottom=Side(style="thin", color="BFBFBF"))

# Formato monetario: negativos entre paréntesis, cero como guión.
_FMT_ARS = "#,##0.00;(#,##0.00);\"-\""
_FMT_ENTERO = "#,##0;(#,##0);\"-\""
_FMT_FECHA = "dd/mm/yyyy"


def _escribir_tabla(ws, df: pd.DataFrame, fila: int, titulo: str | None = None,
                    formatos: dict[str, str] | None = None) -> int:
    """Escribe un DataFrame como tabla formateada; devuelve la fila siguiente."""
    formatos = formatos or {}
    if titulo:
        ws.cell(row=fila, column=1, value=titulo).font = _FONT_BOLD
        fila += 1
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=fila, column=j, value=str(col))
        c.font = _FONT_BOLD
        c.fill = _FILL_HEADER
        c.border = _BORDE
    fila += 1
    for _, row in df.iterrows():
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = None
            elif isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            c = ws.cell(row=fila, column=j, value=val)
            c.font = _FONT
            fmt = formatos.get(str(col))
            if fmt:
                c.number_format = fmt
        fila += 1
    return fila + 1


def _ancho_columnas(ws, anchos: dict[int, int]) -> None:
    for idx, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(idx)].width = ancho


def _autoancho(ws, df: pd.DataFrame, maximo: int = 55) -> None:
    for j, col in enumerate(df.columns, start=1):
        contenido = df[col].astype(str).str.len().max() if len(df) else 0
        ws.column_dimensions[get_column_letter(j)].width = min(
            max(int(contenido or 0) + 2, len(str(col)) + 2), maximo
        )


def generar_excel(
    path: str | Path,
    meta: dict,
    arqueo: pd.DataFrame,
    medios: pd.DataFrame,
    bolsines: pd.DataFrame,
    difc_eventos: pd.DataFrame,
    difc_resumen: pd.DataFrame,
    alertas: pd.DataFrame,
    cascada: pd.DataFrame,
    usd: pd.DataFrame,
) -> Path:
    """Escribe el workbook completo y devuelve el path."""
    path = Path(path)
    wb = Workbook()

    # --- Resumen ---------------------------------------------------------
    ws = wb.active
    ws.title = "Resumen"
    ws.cell(row=1, column=1, value="ARQUEO DE CAJA — MORENO").font = _FONT_TITULO
    ws.cell(row=2, column=1, value=f"Período: {meta['desde']:%d/%m/%Y} al {meta['hasta']:%d/%m/%Y}").font = _FONT
    ws.cell(row=3, column=1, value=f"Empresas del export: {meta.get('empresas', 's/d')} "
            "(⚠ el export no distingue empresa por asiento)").font = _FONT
    ws.cell(row=4, column=1, value=f"Generado: {meta['generado']}").font = _FONT

    dif_total = float(difc_resumen["neto_final"].sum()) if len(difc_resumen) else 0.0
    n_rojas = int((alertas["severidad"] == "ROJA").sum()) if len(alertas) else 0
    c = ws.cell(row=6, column=1, value="Diferencia neta de la ventana (post-correcciones):")
    c.font = _FONT_BOLD
    c = ws.cell(row=6, column=4, value=dif_total)
    c.font = _FONT_BOLD
    c.number_format = _FMT_ARS
    c = ws.cell(row=7, column=1, value="Alertas ROJAS abiertas:")
    c.font = _FONT_BOLD
    ws.cell(row=7, column=4, value=n_rojas).font = _FONT_BOLD

    fila = 9
    medios_fmt = {c: _FMT_ENTERO for c in
                  ["efectivo", "mercado_pago", "tarjetas", "cheques", "total"]}
    medios_fmt["fecha"] = _FMT_FECHA
    fila = _escribir_tabla(ws, medios, fila, "Cobranza diaria por medio de pago (ARS)", medios_fmt)

    if len(arqueo):
        por_caja = (
            arqueo.groupby("caja")
            .agg(
                dias=("fecha", "nunique"),
                cobrado_efectivo=("cobrado_efectivo", "sum"),
                diferencia_registrada=("diferencia_registrada", "sum"),
                dias_revisar=("estado", lambda s: int((s == "REVISAR").sum())),
            )
            .reset_index()
        )
        fila = _escribir_tabla(
            ws, por_caja, fila, "Totales de la ventana por caja (ARS)",
            {"cobrado_efectivo": _FMT_ENTERO, "diferencia_registrada": _FMT_ARS},
        )
    _ancho_columnas(ws, {1: 26, 2: 16, 3: 16, 4: 18, 5: 16, 6: 14})

    # --- Arqueo diario ----------------------------------------------------
    ws = wb.create_sheet("Arqueo diario")
    cols_arqueo = ["fecha", "caja", "cajera", "apertura", "cobrado_efectivo",
                   "vueltos", "alivios", "cierres", "diferencia_registrada",
                   "otros", "neto_dia", "n_cobranzas", "estado"]
    tabla = arqueo[cols_arqueo] if len(arqueo) else pd.DataFrame(columns=cols_arqueo)
    fmt = {c: _FMT_ARS for c in cols_arqueo[3:11]}
    fmt.update({"fecha": _FMT_FECHA, "n_cobranzas": _FMT_ENTERO})
    _escribir_tabla(
        ws, tabla, 1,
        "Descomposición diaria por caja — apertura + cobrado − vueltos − alivios "
        "− cierres + diferencia + otros = neto (debe dar 0)",
        fmt,
    )
    _autoancho(ws, tabla)
    ws.freeze_panes = "A3"

    # --- Bolsines -----------------------------------------------------------
    ws = wb.create_sheet("Bolsines")
    fmt = {"monto_entrada": _FMT_ARS, "monto_salida": _FMT_ARS,
           "diferencia": _FMT_ARS, "fecha_cierre": _FMT_FECHA,
           "fecha_barrido": _FMT_FECHA}
    _escribir_tabla(
        ws, bolsines, 1,
        "Conciliación de bolsines precintados: cierre caja→buzón vs barrido "
        "buzón→puente (por bolsa+precinto)",
        fmt,
    )
    _autoancho(ws, bolsines)
    ws.freeze_panes = "A3"

    # --- Diferencias ----------------------------------------------------------
    ws = wb.create_sheet("Diferencias")
    fila = _escribir_tabla(
        ws, difc_resumen, 1,
        "Diferencia FINAL por caja (neteadas las reversiones de tesorería)",
        {"neto_final": _FMT_ARS, "bruto_movido": _FMT_ARS},
    )
    ev = difc_eventos.drop(columns=["cuenta_id"], errors="ignore")
    fmt = {"efecto": _FMT_ARS, "fecha": _FMT_FECHA,
           "ingreso": "dd/mm/yyyy hh:mm"}
    _escribir_tabla(ws, ev, fila, "Todos los asientos DIFC de la ventana", fmt)
    _autoancho(ws, ev)

    # --- Alertas ------------------------------------------------------------
    ws = wb.create_sheet("Alertas")
    _escribir_tabla(ws, alertas, 1, "Alertas (ROJA = actuar hoy)", {
        "monto": _FMT_ARS, "fecha": _FMT_FECHA,
    })
    for row in ws.iter_rows(min_row=2, max_col=1):
        for c in row:
            if c.value == "ROJA":
                c.font = _FONT_BOLD
    _autoancho(ws, alertas)
    ws.freeze_panes = "A3"

    # --- Cascada -------------------------------------------------------------
    ws = wb.create_sheet("Cascada")
    fmt = {"entradas": _FMT_ENTERO, "salidas": _FMT_ENTERO,
           "neto_dia": _FMT_ENTERO, "saldo_relativo": _FMT_ENTERO,
           "fecha": _FMT_FECHA}
    _escribir_tabla(
        ws, cascada.drop(columns=["cuenta_id"], errors="ignore"), 1,
        "Cascada de consolidación (buzón → puente → fuerte). OJO: el saldo es "
        "RELATIVO al inicio de la ventana, no el efectivo absoluto.",
        fmt,
    )
    _autoancho(ws, cascada)

    # --- Caja USD --------------------------------------------------------------
    ws = wb.create_sheet("Caja USD")
    fmt = {"usd": "#,##0.00", "ars": _FMT_ARS, "cotizacion": "#,##0.00",
           "saldo_usd_relativo": "#,##0.00", "fecha": _FMT_FECHA}
    _escribir_tabla(
        ws, usd, 1,
        "Caja dólar (arqueada en USD por columna Nominal; saldo relativo a la ventana)",
        fmt,
    )
    _autoancho(ws, usd)

    # --- Metodología --------------------------------------------------------------
    ws = wb.create_sheet("Metodología")
    lineas = [
        ("CÓMO LEER ESTE ARQUEO", True),
        ("", False),
        ("1. 'Arqueo diario': cada caja de venta debe cerrar el día en neto $0. La", False),
        ("   identidad es: apertura + cobrado − vueltos − alivios − cierres +", False),
        ("   diferencia registrada + otros = neto. Un neto distinto de cero significa", False),
        ("   día contablemente incompleto, no necesariamente plata faltante.", False),
        ("2. 'Bolsines': cada bolsa precintada se asienta dos veces (cierre caja→buzón", False),
        ("   y barrido buzón→puente a la mañana siguiente). Acá se concilian por", False),
        ("   bolsa+precinto. PENDIENTE en el último día de la ventana es normal.", False),
        ("3. 'Diferencias': los asientos DIFC con sus reversiones. Lo que importa es el", False),
        ("   neto FINAL por caja; el bruto movido mide el ruido del proceso de conteo.", False),
        ("4. 'Cascada': flujos del buzón/puente/caja fuerte. Los saldos son relativos al", False),
        ("   inicio de la ventana porque el export no trae saldos iniciales.", False),
        ("", False),
        ("LÍMITES CONOCIDOS DEL EXPORT", True),
        ("- No distingue empresa (0008/0009) por asiento: el arqueo es del consolidado.", False),
        ("- Un asiento borrado en Sigma no desaparece del snapshot local.", False),
        ("- El export solo muestra la última modificación, no qué cambió.", False),
        ("- Sugerencia operativa: exportar el diario cada mañana con 7 días hacia atrás,", False),
        ("  así las correcciones tardías (hasta 3 días) quedan siempre incluidas.", False),
        ("", False),
        ("Razonamiento completo: docs/arqueo/README.md del repo masmelos-analytics.", False),
    ]
    for i, (texto, negrita) in enumerate(lineas, start=1):
        c = ws.cell(row=i, column=1, value=texto)
        c.font = _FONT_BOLD if negrita else _FONT
    ws.column_dimensions["A"].width = 95

    for hoja in wb.worksheets:
        hoja.sheet_view.showGridLines = True
        hoja.page_setup.orientation = "landscape"
        hoja.page_setup.fitToWidth = 1

    path.parent.mkdir(parents=True, exist_ok=True)
    destino = _guardar_sin_pisar_abierto(wb, path)
    logger.info("Excel de arqueo escrito en %s", destino)
    return destino


def _guardar_sin_pisar_abierto(wb, path: Path) -> Path:
    """Guarda el workbook; si el destino está abierto en Excel (lockeado),
    escribe una copia con sufijo en vez de crashear.

    En la operatoria diaria es normal que el tesorero deje abierto el informe
    del día y vuelva a correr el arqueo (p. ej. tras cargar una corrección en
    Sigma). Sin esto, la corrida moría con PermissionError.
    """
    try:
        wb.save(path)
        return path
    except PermissionError:
        for i in range(2, 100):
            alt = path.with_name(f"{path.stem}_{i}{path.suffix}")
            try:
                wb.save(alt)
                logger.warning(
                    "%s estaba abierto en Excel; escribí la copia %s "
                    "(cerrá el original para que la próxima lo actualice).",
                    path.name, alt.name)
                return alt
            except PermissionError:
                continue
        raise
